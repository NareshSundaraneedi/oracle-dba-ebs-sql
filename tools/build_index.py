#!/usr/bin/env python3
"""Build SQL_Toolkit_Index.xlsx and MASTER_INDEX.md from generated SQL headers."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
TK = ROOT / "Oracle_DBA_SQL_Toolkit"
XLSX = ROOT / "SQL_Toolkit_Index.xlsx"
MD = TK / "MASTER_INDEX.md"

HDR = re.compile(r"^--\s+(File Name|Category|Purpose|Oracle Version|EBS Version|Difficulty|Production Use)\s*:\s*(.*)$")


def parse_sql(path: Path) -> dict:
    meta = {
        "File Name": path.name,
        "Category": path.parent.name,
        "Purpose": "",
        "Oracle Version": "19c",
        "EBS Version": "",
        "Difficulty": "",
        "Production Use": "YES",
    }
    text = path.read_text(encoding="utf-8", errors="replace").splitlines()
    for line in text[:25]:
        m = HDR.match(line)
        if m:
            meta[m.group(1)] = m.group(2).strip()
    # privileges from first query block
    priv = "SELECT_CATALOG_ROLE"
    for line in text:
        if "Required privileges" in line or "Required Privileges" in line:
            continue
        if line.startswith("--    ") and "SELECT" in line.upper() and "privileges" not in line.lower():
            pass
        if "Required privileges" in line.lower() or line.strip().startswith("--    SELECT") or line.strip().startswith("--    APPS") or line.strip().startswith("--    SYS") or line.strip().startswith("--    AUDIT"):
            val = line.split(":", 1)[-1].strip() if ":" in line else line.replace("--", "").strip()
            if val and val not in ("Required privileges",):
                if len(val) > 3 and not val.startswith("1."):
                    priv = val[:80]
                    break
    for i, line in enumerate(text):
        if "7. Required privileges" in line and i + 1 < len(text):
            priv = text[i + 1].lstrip("- ").strip()[:100] or priv
            break
    ebs = "N/A"
    ev = meta.get("EBS Version", "")
    if "12.2" in ev or "R12" in ev:
        ebs = "R12.2"
    if path.parent.name.startswith("2") or "EBS" in path.parent.name:
        ebs = "R12.2"
    return {
        "Category": meta["Category"],
        "Script Name": meta["File Name"],
        "Purpose": meta["Purpose"] or path.stem.replace("_", " "),
        "Difficulty": meta["Difficulty"] or "Intermediate",
        "Production Safe": meta["Production Use"] or "YES",
        "Required Privilege": priv,
        "Oracle Version": meta["Oracle Version"] or "19c",
        "EBS": ebs,
        "Folder": path.parent.name,
        "RelPath": str(path.relative_to(ROOT)),
    }


def main() -> None:
    rows = []
    for p in sorted(TK.rglob("*.sql")):
        rows.append(parse_sql(p))

    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Toolkit Index"
    headers = [
        "Category",
        "Script Name",
        "Purpose",
        "Difficulty",
        "Production Safe",
        "Required Privilege",
        "Oracle Version",
        "EBS",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    wrap = Alignment(wrap_text=True, vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = thin
    fills = {
        "Basic": PatternFill("solid", fgColor="E2EFDA"),
        "Intermediate": PatternFill("solid", fgColor="FFF2CC"),
        "Advanced": PatternFill("solid", fgColor="FCE4D6"),
    }
    for r, row in enumerate(rows, 2):
        vals = [row[h] for h in headers]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = wrap
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
        if row["Difficulty"] in fills:
            ws.cell(r, 4).fill = fills[row["Difficulty"]]
        if str(row["Production Safe"]).upper().startswith("YES"):
            ws.cell(r, 5).fill = PatternFill("solid", fgColor="C6EFCE")
    ws.auto_filter.ref = f"A1:H{len(rows)+1}"
    ws.freeze_panes = "A2"
    widths = [28, 46, 72, 14, 16, 42, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    # Summary sheet
    sm = wb.create_sheet("Summary", 0)
    sm["A1"] = "Oracle DBA + EBS DBA SQL Toolkit — Index Summary"
    sm["A1"].font = Font(bold=True, size=16, color="1F4E79")
    sm["A3"] = "Total scripts"
    sm["B3"] = len(rows)
    sm["A4"] = "Oracle version"
    sm["B4"] = "19c"
    sm["A5"] = "EBS version"
    sm["B5"] = "R12.2.x"
    sm["A7"] = "Category"
    sm["B7"] = "Scripts"
    sm["A7"].font = Font(bold=True)
    sm["B7"].font = Font(bold=True)
    from collections import Counter

    c = Counter(r["Category"] for r in rows)
    for i, (cat, n) in enumerate(sorted(c.items()), 8):
        sm.cell(i, 1, cat)
        sm.cell(i, 2, n)
    sm.column_dimensions["A"].width = 40
    sm.column_dimensions["B"].width = 14

    lic = wb.create_sheet("Licensing")
    lic["A1"] = "Licensing notes"
    lic["A1"].font = Font(bold=True, size=14)
    notes = [
        "V$SESSION, V$SQL, V$SQL_PLAN, V$SYSSTAT, DBA_* dictionary views: included with the database (EE features vary).",
        "V$ACTIVE_SESSION_HISTORY, DBA_HIST_*: Oracle Diagnostics Pack.",
        "V$SQL_MONITOR, DBMS_SQLTUNE / SQL Tuning Advisor, SQL profiles via advisor: Tuning Pack.",
        "AWR reports (awrrpt.sql / DISPLAY_AWR): Diagnostics Pack.",
        "SQL Plan Management baselines: Enterprise Edition.",
        "Unified audit views: 19c database feature; Fine Grained Auditing is EE.",
        "Privilege Analysis / Database Vault / In-Memory: respective options.",
        "Scripts that need a pack say so in the file header extra notes and query comments.",
        "This index is a catalog, not a license audit.",
    ]
    for i, n in enumerate(notes, 3):
        lic.cell(i, 1, n)
        lic.cell(i, 1).alignment = Alignment(wrap_text=True)
    lic.column_dimensions["A"].width = 120

    wb.save(XLSX)

    # Markdown master index
    lines = [
        "# SQL Toolkit Master Index",
        "",
        f"Total scripts: **{len(rows)}**  ",
        "Oracle: **19c** · EBS: **R12.2.x** · RAC / ASM / Data Guard scripts are included and no-op safely when those features are absent.",
        "",
        "File naming: `NN_short_snake_case.sql` inside each category folder.",
        "",
    ]
    current = None
    for r in rows:
        if r["Category"] != current:
            current = r["Category"]
            lines.append(f"## {current}")
            lines.append("")
            lines.append("| Script | Purpose | Difficulty | Production | EBS |")
            lines.append("|---|---|---|---|---|")
        lines.append(
            f"| `{r['Script Name']}` | {r['Purpose']} | {r['Difficulty']} | {r['Production Safe']} | {r['EBS']} |"
        )
    lines.append("")
    MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {XLSX} ({len(rows)} rows)")
    print(f"Wrote {MD}")


if __name__ == "__main__":
    main()
